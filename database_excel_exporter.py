import os
from dotenv import load_dotenv
from datetime import datetime
import psycopg2
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font


class DatabaseExcelExporter:
    def __init__(self):
        # Load environment variables from .env local file
        load_dotenv()
        # Connect to the database
        self.conn = psycopg2.connect(
            host=os.getenv("YOUR_DATABASE_HOST"),
            port=os.getenv("YOUR_DATABASE_PORT"),
            user=os.getenv("YOUR_DATABASE_USER"),
            password=os.getenv("YOUR_DATABASE_PASSWORD"),
            database=os.getenv("YOUR_DATABASE_NAME")
        )
        self.workbook = Workbook()

    def export_tables(self):

        cur = self.conn.cursor()
        cur.execute("SELECT table_name FROM information_schema.tables WHERE table_schema = %s", (os.getenv("YOUR_DATABASE_SCHEMA_NAME"),))
        table_names = [table[0] for table in cur.fetchall()]

        for table_name in table_names:
            self.export_table(table_name)

        self.workbook.save(f'{os.getenv("YOUR_FILENAME")}_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx')
        self.conn.close()

    def export_table(self, table_name):
        cur = self.conn.cursor()
        cur.execute(f"SELECT * FROM {os.getenv('YOUR_DATABASE_SCHEMA_NAME')}.{table_name}")
        headers = [column[0] for column in cur.description]

        worksheet = self.workbook.create_sheet(title=table_name)
        self.write_headers(worksheet, headers)
        self.write_data(worksheet, cur)
        self.autofit_columns(worksheet)

    def write_headers(self, worksheet, headers):
        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

    def write_data(self, worksheet, cur):
        for row_num, row in enumerate(cur, start=2):
            for col_num, cell_value in enumerate(row, start=1):
                if isinstance(cell_value, int):
                    worksheet.cell(row=row_num, column=col_num).value = cell_value
                else:
                    worksheet.cell(row=row_num, column=col_num).value = str(cell_value)

    def autofit_columns(self, worksheet):
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width
